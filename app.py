import streamlit as st
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font as xlFont, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape as rl_landscape, portrait as rl_portrait
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="V.A.G.A.L.U.M.E. Pro", layout="wide")

# --- FUNÇÃO DE FONTES (À PROVA DE ERROS) ---
def inicializar_fontes():
    # Fontes nativas que nunca dão erro
    fontes = ["Helvetica", "Times-Roman", "Courier"]
    
    # Verifica se o arquivo TTF existe na pasta (Você pode subir o .ttf no GitHub)
    # Se não existir, ele ignora e não quebra o app
    if os.path.exists("Roboto-Regular.ttf"):
        try:
            pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Regular.ttf'))
            fontes.append("Roboto")
        except:
            pass
    return fontes

# Inicializa a lista de fontes uma única vez
if 'fontes_lista' not in st.session_state:
    st.session_state.fontes_lista = inicializar_fontes()

# --- INTERFACE ---
st.title("💡 V.A.G.A.L.U.M.E. - Sistema de Relatórios")

st.sidebar.header("⚙️ Configurações")
ano_sel = st.sidebar.number_input("Ano do Relatório:", value=datetime.now().year)
fonte_sel = st.sidebar.selectbox("Escolha a Fonte", st.session_state.fontes_lista)

with st.sidebar.expander("🎨 Cores e Estilo"):
    cor_rota = st.sidebar.color_picker("Cor das Rotas", "#1E3A8A")
    cor_txt_rota = st.sidebar.color_picker("Texto das Rotas", "#FFFFFF")
    cor_bairro = st.sidebar.color_picker("Cor dos Bairros", "#D3D3D3")

# Mapeamento de Colunas (Fixo conforme sua planilha)
COL_DATA = 7      # Coluna H
COL_PROBLEMA = 1  # Coluna B
COL_STATUS = 3    # Coluna D

# Dicionário de Rotas
routes = {
    "ROTA 1": ["CENTRO", "JARDIM AMERICA"],
    "ROTA 2": ["ALBERTINA", "LARANJEIRAS", "BOA VISTA", "EUGENIO SCHNEIDER"],
    "ROTA 3": ["FUNDO CANOAS", "CANOAS", "PROGRESSO", "PAMPLONA", "CANTA GALO"],
    "ROTA 4": ["BARRA DO TROMBUDO", "BARRAGEM", "BUDAG", "SUMARE"],
    "ROTA 5": ["SANTANA", "TABOAO", "BREMER", "BELA ALIANÇA"],
    "ROTA 6": ["BARRA DA ITOUPAVA", "NAVEGANTES", "SANTA RITA", "VALADA ITOUPAVA", "VALADA SÃO PAULO", "RAINHA"]
}

uploaded_file = st.file_uploader("📥 Envie a planilha Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    if st.button("🚀 Gerar Relatórios"):
        with st.spinner('Processando dados da Coluna H...'):
            try:
                xls = pd.read_excel(uploaded_file, sheet_name=None, header=None)
                abas_disponiveis = {nome.strip().upper(): nome for nome in xls.keys()}
                
                data_by_route = {}
                contador = 0

                for route, neighborhoods in routes.items():
                    route_data = {}
                    for neighborhood in neighborhoods:
                        nome_upper = neighborhood.upper()
                        if nome_upper in abas_disponiveis:
                            df = xls[abas_disponiveis[nome_upper]].copy()
                            if df.shape[1] <= COL_DATA: continue

                            # Filtro de Data e Status
                            df[COL_DATA] = pd.to_datetime(df[COL_DATA], errors='coerce')
                            mask = (
                                (df[COL_DATA].dt.year == ano_sel) & 
                                (df[COL_STATUS].astype(str).str.strip().str.upper().isin(['NÃO REALIZADO', 'NÃO EXECUTADO', 'NAO REALIZADO', 'NAO EXECUTADO']))
                            )
                            
                            items = df[mask][COL_PROBLEMA].dropna().astype(str).str.strip().tolist()
                            if items:
                                route_data[neighborhood] = items
                                contador += len(items)
                    
                    if route_data:
                        data_by_route[route] = route_data

                if contador == 0:
                    st.warning(f"Nenhum dado de {ano_sel} encontrado.")
                else:
                    # --- GERAÇÃO EXCEL ---
                    ex_buffer = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    row = 1
                    for r, bairros in data_by_route.items():
                        c = ws.cell(row=row, column=1, value=r)
                        c.font = xlFont(bold=True, color=cor_txt_rota.replace('#',''))
                        c.fill = PatternFill(start_color=cor_rota.replace('#',''), fill_type="solid")
                        row += 1
                        for b, probs in bairros.items():
                            ws.cell(row=row, column=1, value=f"📍 {b}").font = xlFont(bold=True)
                            row += 1
                            for p in probs:
                                ws.cell(row=row, column=1, value=f"  - {p}")
                                row += 1
                    wb.save(ex_buffer)
                    ex_buffer.seek(0)

                    # --- GERAÇÃO PDF ---
                    pdf_buffer = io.BytesIO()
                    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
                    
                    # Define negrito com base na fonte escolhida
                    f_bold = "Helvetica-Bold"
                    if fonte_sel == "Roboto": f_bold = "Roboto-Bold"
                    elif fonte_sel == "Times-Roman": f_bold = "Times-Bold"
                    
                    style_r = ParagraphStyle('R', fontName=f_bold, fontSize=14, textColor=HexColor(cor_txt_rota), backColor=HexColor(cor_rota), alignment=1, padding=5)
                    style_b = ParagraphStyle('B', fontName=f_bold, fontSize=11, leftIndent=10, backColor=HexColor(cor_bairro))
                    style_i = ParagraphStyle('I', fontName=fonte_sel, fontSize=10, leftIndent=20)

                    elements = [Paragraph(f"RELATÓRIO DE PENDÊNCIAS - {ano_sel}", style_r), Spacer(1, 15)]
                    for r, bairros in data_by_route.items():
                        elements.append(Paragraph(r, style_r))
                        for b, probs in bairros.items():
                            elements.append(Paragraph(b, style_b))
                            for p in probs:
                                elements.append(Paragraph(f"• {p}", style_i))
                        elements.append(Spacer(1, 10))

                    doc.build(elements)
                    pdf_buffer.seek(0)

                    st.success(f"Sucesso! {contador} itens encontrados.")
                    st.download_button("📥 Baixar Excel", data=ex_buffer, file_name=f"Vagalume_{ano_sel}.xlsx")
                    st.download_button("📄 Baixar PDF", data=pdf_buffer, file_name=f"Vagalume_{ano_sel}.pdf")

            except Exception as e:
                st.error(f"Erro inesperado: {e}")
            
